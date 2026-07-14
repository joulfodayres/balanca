import os
import io
import psycopg2
import psycopg2.extras
from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
CORS(app)

DATABASE_URL = os.environ.get("DATABASE_URL")


def get_db():
    return psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)


def init_db():
    if not DATABASE_URL:
        return
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS registos (
            id   SERIAL PRIMARY KEY,
            data DATE    NOT NULL UNIQUE,
            peso REAL    NOT NULL
        )
    """)
    conn.commit()
    cur.close()
    conn.close()


init_db()


# ─── Frontend ────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


# ─── API: Registos ────────────────────────────────────────────────────────────

@app.route("/api/registos", methods=["GET"])
def list_registos():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, data::text, peso FROM registos ORDER BY data ASC")
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify(rows)


@app.route("/api/registos", methods=["POST"])
def create_registo():
    body = request.get_json()
    data = body.get("data")
    peso = body.get("peso")
    if not data or peso is None:
        return jsonify({"erro": "data e peso são obrigatórios"}), 400
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO registos (data, peso) VALUES (%s, %s)", (data, peso)
        )
        conn.commit()
    except psycopg2.IntegrityError:
        conn.rollback()
        cur.close()
        conn.close()
        return jsonify({"erro": f"Já existe registo para {data}"}), 409
    cur.close()
    conn.close()
    return jsonify({"ok": True}), 201


@app.route("/api/registos/<int:rid>", methods=["PUT"])
def update_registo(rid):
    body = request.get_json()
    data = body.get("data")
    peso = body.get("peso")
    if not data or peso is None:
        return jsonify({"erro": "data e peso são obrigatórios"}), 400
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "UPDATE registos SET data=%s, peso=%s WHERE id=%s", (data, peso, rid)
    )
    rowcount = cur.rowcount
    conn.commit()
    cur.close()
    conn.close()
    if rowcount == 0:
        return jsonify({"erro": "Registo não encontrado"}), 404
    return jsonify({"ok": True})


@app.route("/api/registos/<int:rid>", methods=["DELETE"])
def delete_registo(rid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM registos WHERE id=%s", (rid,))
    rowcount = cur.rowcount
    conn.commit()
    cur.close()
    conn.close()
    if rowcount == 0:
        return jsonify({"erro": "Registo não encontrado"}), 404
    return jsonify({"ok": True})


# ─── API: Import bulk ─────────────────────────────────────────────────────────

@app.route("/api/import", methods=["POST"])
def import_registos():
    items = request.get_json()
    if not isinstance(items, list):
        return jsonify({"erro": "Esperado array de {data, peso}"}), 400
    inserted = 0
    skipped = 0
    conn = get_db()
    cur = conn.cursor()
    for item in items:
        try:
            cur.execute(
                "INSERT INTO registos (data, peso) VALUES (%s, %s) ON CONFLICT (data) DO NOTHING",
                (item["data"], item["peso"]),
            )
            if cur.rowcount > 0:
                inserted += 1
            else:
                skipped += 1
        except Exception:
            skipped += 1
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"inseridos": inserted, "ignorados": skipped})


# ─── API: Pesquisa histórica ──────────────────────────────────────────────────

@app.route("/api/pesquisa", methods=["GET"])
def pesquisa():
    data = request.args.get("data")
    if not data:
        return jsonify({"erro": "Parâmetro 'data' obrigatório"}), 400
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT peso FROM registos WHERE data = %s", (data,))
    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        return jsonify({"erro": f"Sem registo para {data}"}), 404
    peso_ref = row["peso"]
    cur.execute(
        """
        SELECT data::text, peso FROM registos
        WHERE peso <= %s AND data <= %s
        ORDER BY data DESC
        LIMIT 10
        """,
        (peso_ref, data),
    )
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify({
        "data": data,
        "peso_ref": peso_ref,
        "resultados": rows,
    })


# ─── API: Exportar Excel ──────────────────────────────────────────────────────

@app.route("/api/export", methods=["GET"])
def export_excel():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT data::text, peso FROM registos ORDER BY data ASC")
    rows = cur.fetchall()
    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Registos"

    # Cabeçalho
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    for col, title in enumerate(["Data", "Peso (kg)"], start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12

    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=row["data"])
        ws.cell(row=i, column=2, value=round(row["peso"], 1))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="balanca.xlsx",
    )


# ─── API: Pesquisa por peso ───────────────────────────────────────────────────

@app.route("/api/pesquisa-peso", methods=["GET"])
def pesquisa_peso():
    peso_str = request.args.get("peso")
    if not peso_str:
        return jsonify({"erro": "Parâmetro 'peso' obrigatório"}), 400
    try:
        peso = float(peso_str)
    except ValueError:
        return jsonify({"erro": "Peso inválido"}), 400
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT data::text, peso FROM registos
        WHERE peso <= %s
        ORDER BY data DESC
        LIMIT 10
        """,
        (peso,),
    )
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify({"peso_ref": peso, "resultados": rows})


# ─── API: Apagar todos os dados ───────────────────────────────────────────────

@app.route("/api/apagar-tudo", methods=["DELETE"])
def apagar_tudo():
    body = request.get_json()
    if not body or body.get("password") != "1973":
        return jsonify({"erro": "Password incorreta"}), 403
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM registos")
    deleted = cur.rowcount
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"ok": True, "eliminados": deleted})


if __name__ == "__main__":
    app.run(debug=True)
